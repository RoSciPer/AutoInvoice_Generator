import unicodedata
import re
import os
import json
import glob
import openpyxl
import pdfplumber
import tkinter as tk
from tkinter import messagebox
from datetime import datetime
import tkinter as tk
from tkinter import messagebox
from help import nolasit_rekina_numuru, saglabat_rekina_numuru

def padari_faila_nosaukumu_drosu(nosaukums):
    # Aizvieto Windows nelegālos simbolus ar apakšsvītru
    nosaukums = re.sub(r'[<>:"/\\|?*\n\r]', '_', nosaukums)
    return nosaukums.strip()

def noņem_diakritiku(teksts):
    # Noņem diakritiskās zīmes, piemēram, ž, ā, ē utt.
    return ''.join(
        c for c in unicodedata.normalize('NFD', teksts)
        if unicodedata.category(c) != 'Mn'
    )

# Funkcija, lai nolasītu konfigurācijas failu
def nolasit_konfiguraciju():
    try:
        with open("config.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except UnicodeDecodeError as e:
        print(f"Kļūda, nolasot konfigurācijas failu: {e}")
        raise
    
    """
    Nolasa konfigurācijas failu un atgriež datus kā vārdnīcu.
    Ja fails neeksistē vai ir bojāts, atgriež noklusējuma vērtības.
    """
    nokluseta_konfiguracija = {
        "pdf_directory": "C:/Users/janis/Documents/rekinsuzzina",
        "save_directory": "C:/Users/janis/Documents/rekini"
    }

    # Pārbauda, vai konfigurācijas fails eksistē
    if os.path.exists(fails):
        try:
            with open(fails, "r") as f:
                return json.load(f)
        except json.JSONDecodeError:
            print("Kļūda: Konfigurācijas fails ir bojāts. Tiek izmantotas noklusējuma vērtības.")
    return nokluseta_konfiguracija

# Funkcija, lai saglabātu konfigurācijas failu
def saglabat_konfiguraciju(config, fails="config.json"):
    """
    Saglabā datus konfigurācijas failā.
    """
    with open(fails, "w") as f:
        json.dump(config, f, indent=4)

def formatet_latviesu_dienas_datumu(datums=None):
    if datums is None:
        datums = datetime.today()
    menesi = [
        "janvāris", "februāris", "marts", "aprīlis", "maijs", "jūnijs",
        "jūlijs", "augusts", "septembris", "oktobris", "novembris", "decembris"
    ]
    return f"{datums.year}. gada {datums.day}. {menesi[datums.month - 1]}"


# Vieninieki
vieninieki = ["", "viens", "divi", "trīs", "četri", "pieci", "seši", "septiņi", "astoņi", "deviņi"]

# Simti
simti = [f"{v} simti" if i != 1 else "simts" for i, v in enumerate(vieninieki)]

# Pusaudži (11-19)
pusaudzi = ["desmit", "vienpadsmit", "divpadsmit", "trīspadsmit", "četrpadsmit",
            "piecpadsmit", "sešpadsmit", "septiņpadsmit", "astoņpadsmit", "deviņpadsmit"]

# Desmitnieki (20, 30, ..., 90)
desmitnieki = [
    "", "desmit", "divdesmit", "trīsdesmit", "četrdesmit",
    "piecdesmit", "sešdesmit", "septiņdesmit", "astoņdesmit", "deviņdesmit"
]

# Funkcija, lai pārvērstu skaitli vārdos
def triju_ciparu_vārdi(skaitlis):
    """
    Funkcija, kas pārvērš trīs ciparu skaitli vārdos latviešu valodā.
    """
    skaitlis = int(skaitlis)
    vārdi = []
    
    # Simti
    if skaitlis >= 100:
        vārdi.append(simti[skaitlis // 100])
        skaitlis %= 100
    
    # Pusaudži (11-19)
    if 10 < skaitlis < 20:
        vārdi.append(pusaudzi[skaitlis - 10])
        return " ".join(vārdi)
    
    # Desmitnieki
    if skaitlis >= 10:
        vārdi.append(desmitnieki[skaitlis // 10])
        skaitlis %= 10
    
    # Vieninieki
    if skaitlis > 0:
        vārdi.append(vieninieki[skaitlis])
    
    return " ".join(vārdi)


def number_to_words_lv(skaidra_summa):
    """
    Funkcija, kas konvertē summu uz vārdiem latviešu valodā un novērš dubultu "eiro" pievienošanu.
    """
    skaidra_summa = f"{float(skaidra_summa):.2f}"
    vesela_daļa, decimāldaļa = map(int, skaidra_summa.split("."))
    summas_vārdi = []
    
    # Tūkstoši
    if vesela_daļa >= 1000:
        tukstosa_daļa = vesela_daļa // 1000
        summas_vārdi.append(f"{triju_ciparu_vārdi(tukstosa_daļa)} tūkstoši")
        vesela_daļa %= 1000  # Atlikums
    
    # Simti, desmitnieki, vieninieki
    if vesela_daļa > 0:
        summas_vārdi.append(triju_ciparu_vārdi(vesela_daļa))
    
    # Pievieno "eiro" tikai vienu reizi
    summas_vārdi.append("eiro")
    
    # Centi
    centu_vārdi = f"{decimāldaļa} centi"  # Centus attēlo kā "23 centi"
    
    # Apvieno tekstu ar "un" tikai starp pēdējām divām daļām (eiro un centi)
    return f"{' '.join(summas_vārdi)} un {centu_vārdi}".capitalize()

# Funkcija, lai atrastu pēdējo modificēto PDF failu
def atrast_pedejo_pdf(direktorija):
    """
    Atrod pēdējo modificēto PDF failu norādītajā direktorijā.
    """
    pdf_faili = glob.glob(os.path.join(direktorija, "*.pdf"))
    if not pdf_faili:
        print("Direktorijā nav PDF failu.")
        return None
    return max(pdf_faili, key=os.path.getmtime)


# Funkcija, lai aizpildītu Excel šablonu
def aizpildit_excel_sablonu(datums_str, rekina_numurs, klients, personas_kods, summa, sablons, izvada_fails, marka_modelis, ar_pvn, sasijas_nr, aplieciba, degviela, krasa, masa, reg_numurs, apl_numurs):
    """
    Aizpilda Excel šablonu ar datiem un saglabā rēķinu uz norādīto failu.
    """
    wb = openpyxl.load_workbook(sablons)
    ws = wb.active

    # Formatēta summa (ar atstarpēm tūkstošiem, komatu kā decimāldaļu)
    formatēta_summa = f"{float(summa):,.2f}".replace(",", " ").replace(".", ",")

    datums_str = formatet_latviesu_dienas_datumu()
    
    # Piemēram šūnas:
    ws["C1"] = datums_str
    ws["B42"] = datums_str
    ws["D42"] = datums_str
    ws["G1"] = f"TA {rekina_numurs}"


    # --- Pamata informācija ---
    ws["C13"] = klients
    ws["C20"] = klients
    ws["D40"] = klients
    ws["C14"] = personas_kods
    ws["B24"] = f"Vieglā A/M: {marka_modelis}"
    ws["B25"] = f"Šasijas nr. VIN: {sasijas_nr}"
    ws["B26"] = f"Reģ./apl.: {aplieciba}"
    ws["B27"] = f"Degviela: {degviela}"
    ws["B28"] = f"Krāsa: {krasa}"
    ws["B29"] = f"Pašmasa {masa}"
    ws["B30"] = f"Valsts Reģ. Nr. vai Sal. Izziņas Nr.: {reg_numurs}"
    ws["B31"] = f"Īpašum. Apl. Nr.: {apl_numurs}"

    # --- PVN režīms vai peļņas daļas režīms ---
    if ar_pvn:
        ws["A33"] = "PVN 21%"

        neto_summa = float(summa) / 1.21
        pvn_summa = neto_summa * 0.21

        # Šūnas tikai PVN režīmā
        ws["F24"] = f"{neto_summa:.2f}"
        ws["G24"] = f"{neto_summa:.2f}"
        ws["G32"] = f"{neto_summa:.2f}"
        ws["G33"] = f"{pvn_summa:.2f}"
        ws["G34"] = formatēta_summa  # kopējā summa ar PVN

    else:
        ws["A33"] = "Peļņas daļas režīms lietotām precēm PVN likuma 138.pants"

        # Šūnas bez PVN sadalījuma – pilna summa visur
        ws["F24"] = formatēta_summa
        ws["G24"] = formatēta_summa
        ws["G32"] = formatēta_summa
        ws["G34"] = formatēta_summa

    # --- Summa vārdiem ---
    summa_vardiem = number_to_words_lv(summa)
    ws["C36"] = summa_vardiem

    wb.save(izvada_fails)
    print(f"✅ Rēķins saglabāts: {izvada_fails}")


def nolasit_pdf_datus(pdf_fails):
    import pdfplumber

    def atrast_vertibu(rinda, atsl_k):
        atsl_k = atsl_k.lower()
        for i, kol in enumerate(rinda):
            if kol and atsl_k in kol.lower():
                for j in range(i + 1, len(rinda)):
                    if rinda[j] and rinda[j].strip():
                        return rinda[j].strip()
        return ""

    with pdfplumber.open(pdf_fails) as pdf:
        klienta_vards = ""
        personas_kods = ""
        marka_modelis = ""
        sasijas_nr = ""
        aplieciba = ""
        degviela = ""
        krasa = ""
        masa = ""
        reg_numurs = ""
        apl_numurs = ""

        for lapa in pdf.pages:
            tabulas = lapa.extract_tables()
            for tabula in tabulas:
                for rinda in tabula:
                    if not rinda:
                        continue

                    # Klients
                    if "Vārds, uzvārds" in (rinda[0] or ""):
                        klienta_vards = rinda[1].strip() if len(rinda) > 1 and rinda[1] else ""
                        personas_kods = rinda[3].strip() if len(rinda) > 3 and rinda[3] else ""
                        if personas_kods and len(personas_kods) == 11:
                            personas_kods = f"{personas_kods[:6]}-{personas_kods[6:]}"

                    # Marka
                    if "Marka, modelis" in (rinda[0] or ""):
                        marka_modelis = rinda[2].strip() if len(rinda) > 2 and rinda[2] else ""

                    # VIN / Šasijas nr.
                    if "Identifikācijas numurs" in (rinda[0] or "") or "VIN" in (rinda[0] or ""):
                        sasijas_nr = atrast_vertibu(rinda, "VIN")

                    # Reģistrācijas apliecība
                    if "Reģistrācijas apliecības Nr" in (rinda[0] or ""):
                        aplieciba = rinda[1].strip() if len(rinda) > 1 and rinda[1] else ""

                    # VISPĀRĪGIE LAUKI — dinamiskā meklēšana
                    degviela = atrast_vertibu(rinda, "degviela") or degviela
                    krasa = atrast_vertibu(rinda, "krāsa") or krasa
                    masa = atrast_vertibu(rinda, "pašmasa") or masa
                    reg_numurs = atrast_vertibu(rinda, "reģistrācijas numurs") or reg_numurs
                    
                    # Apliecības numuru meklē tikai tekstā (ārpus tabulām)
                    if not apl_numurs:
                        for lapa in pdf.pages:
                            teksts = lapa.extract_text()
                            for rinda in teksts.split('\n'):
                                if "Īpašumtiesību apliecība Nr." in rinda:
                                    apl_numurs = rinda.split("Īpašumtiesību apliecība Nr.")[-1].strip()

        return (
            klienta_vards, personas_kods, marka_modelis,
            sasijas_nr, aplieciba, degviela, krasa,
            masa, reg_numurs, apl_numurs
        )

# Funkcija lai palaistu GUI
def palaist_gui():

    # Nolasām konfigurāciju
    konfiguracija = nolasit_konfiguraciju()

    # Iegūstam direktorijas no konfigurācijas faila
    pdf_direktorija = konfiguracija["pdf_direktorija"]
    saglabāšanas_direktorija = konfiguracija["saglabāšanas_direktorija"]

    global rekina_numura_lauks  # Padarām lauku globālu

    def submit_data():
        summa = ievades_lauks.get()
        ar_pvn = pvn_checkbox.get()

        if not summa:
            messagebox.showerror("Kļūda", "Lūdzu ievadiet summu.")
            return

        try:
            summa = float(summa.replace(",", "."))
        except ValueError:
            messagebox.showerror("Kļūda", "Lūdzu ievadiet derīgu summu.")
            return
        
        # Iegūstam rēķina numuru no ievades lauka
        rekina_numurs = rekina_numura_lauks.get()
        if not rekina_numurs.isdigit():
            messagebox.showerror("Kļūda", "Rēķina numuram jābūt skaitlim.")
            return

        rekina_numurs = int(rekina_numurs)

        # Saglabājam nākamo rēķina numuru
        saglabat_rekina_numuru(rekina_numurs + 1)

        # Pārbaude un direktoriju izmantošana
        if not os.path.exists(pdf_direktorija):
            messagebox.showerror("Kļūda", f"PDF direktorija '{pdf_direktorija}' neeksistē!")
            return

        if not os.path.exists(saglabāšanas_direktorija):
            messagebox.showerror("Kļūda", f"Saglabāšanas direktorija '{saglabāšanas_direktorija}' neeksistē!")
            return
        
        sablons = "Temp.xlsx"
        

        pdf_fails = atrast_pedejo_pdf(pdf_direktorija)
        if not pdf_fails:
            messagebox.showerror("Kļūda", "Neizdevās atrast PDF failu.")
            return

        klients, personas_kods, marka_modelis, sasijas_nr, aplieciba, degviela, krasa, masa, reg_numurs, apl_numurs = nolasit_pdf_datus(pdf_fails)

        faila_nosaukums = f"{marka_modelis.replace(' ', '_')}_{klients.replace(' ', '_')}.xlsx"
        faila_nosaukums = padari_faila_nosaukumu_drosu(noņem_diakritiku(faila_nosaukums))
        izvada_fails = os.path.join(saglabāšanas_direktorija, faila_nosaukums)

        datums_str = formatet_latviesu_dienas_datumu()
        
        aizpildit_excel_sablonu(datums_str, rekina_numurs, klients, personas_kods, summa, sablons, izvada_fails, marka_modelis, ar_pvn, sasijas_nr, aplieciba, degviela, krasa, masa, reg_numurs, apl_numurs)
        messagebox.showinfo("Veiksmīgi", f"Rēķins saglabāts: {izvada_fails}")

        # Automātiski atver failu pēc saglabāšanas
        os.startfile(izvada_fails)

        saknes_logs.after(2000, saknes_logs.destroy)

    # Izveido GUI logu
    saknes_logs = tk.Tk()
    saknes_logs.title("AutoOgre Rēķins")

    # Pievieno ikonu
    saknes_logs.iconbitmap("attels.ico")  # pielāgo ceļu, ja vajag

    # Rēķina numura ievades lauks
    tk.Label(saknes_logs, text="Rēķina numurs:", font=("Arial", 12)).grid(row=0, column=0, padx=10, pady=10)
    rekina_numura_lauks = tk.Entry(saknes_logs, width=20, font=("Arial", 14))
    rekina_numura_lauks.grid(row=0, column=1, padx=10, pady=10)

    # Automātiski aizpilda rēķina numuru
    rekina_numurs = nolasit_rekina_numuru()
    rekina_numura_lauks.insert(0, str(rekina_numurs))

    # Summas ievades lauks
    tk.Label(saknes_logs, text="Ievadiet summu (EUR):", font=("Arial", 12)).grid(row=1, column=0, padx=10, pady=10)
    ievades_lauks = tk.Entry(saknes_logs, width=20, font=("Arial", 14))
    ievades_lauks.grid(row=1, column=1, padx=10, pady=10)
    ievades_lauks.focus_set()

    # PVN izvēles rūtiņa
    pvn_checkbox = tk.BooleanVar()
    tk.Checkbutton(saknes_logs, text="Rēķins ar PVN (21%)", font=("Arial", 12), fg="red", bg="white", variable=pvn_checkbox).grid(row=2, column=0, columnspan=2, padx=10, pady=10)

    # Poga, lai iesniegtu datus
    tk.Button(saknes_logs, text="Izveidot rēķinu", font=("Arial Black", 14), width=30, height=3, command=submit_data).grid(row=3, column=0, columnspan=2, padx=10, pady=10)

    # Enter taustiņš arī izsauc submit_data
    saknes_logs.bind('<Return>', lambda event: submit_data())

    saknes_logs.mainloop()

if __name__ == "__main__":
    palaist_gui()